import re
from pymupdf import pymupdf
from pathlib import Path
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime

BASE_DIR = Path(__file__).parent
PDF_PATH = BASE_DIR / "files" / "BRIAN School Leader Monthly Report 2025-26.pdf"

# Generate versioned output file with timestamp
def get_versioned_output_path():
    """Generate a unique output file path with timestamp."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return BASE_DIR / f"individual_report_data_{timestamp}.xlsx"

OUT_PATH = get_versioned_output_path()

doc =  pymupdf.open(PDF_PATH)

# Pull text from first page of file
page = doc[0]
text = page.get_text()
page_dict = page.get_text("dict")
blocks = page.get_text("blocks")

# Find numbers from document
numbers = re.findall(r"\d+", text)
numbers = [int(n) for n in numbers]
refined_numbers = numbers[37:]

# print(refined_numbers)

# Define School object class
class School:
     def __init__(self, name, is_both, data_list):
          self.name = name
          self.is_both = is_both
          self.data_list = data_list

# Locate school name 
"""""
for block in blocks:
    x0, y0, x1, y1, text_content, block_no, block_type = block
    if block_no == 17: # School name is in block 17
        print(f"Block {block_no}:")
        print(f"  Text: {text_content}")
        print("-" * 20)
"""

# Extract first line of school name block
first_line_text = ""
school_name = ""
if page_dict and "blocks" in page_dict and len(page_dict["blocks"]) > 0:
    target_block = page_dict["blocks"][18] # School name is in block 17

    if "lines" in target_block and len(target_block["lines"]) > 0:
        first_line = target_block["lines"][0] # Get the first line of that block

        # Concatenate spans to get the full line text
        for span in first_line["spans"]:
                first_line_text += span["text"]
    school_name = first_line_text

# Initialize is_both referring to ES and HS
is_both = False
es_and_hs_schools = [
    "Arts and College Preparatory Academy",
    "Columbus Arts and Tech Academy",
    "Columbus Preparatory Academy",
    "Great River Connections Academy",
    "Northeast Ohio College Preparatory School",
    "Ohio Connections Academy",
    "Ohio Virtual Academy",
    "Wildwood Environmental Academy",
    # "Brian's Sample School"
]

# Set is_both to True if school name is found in list
if school_name in es_and_hs_schools:
     is_both = True


# Index label objects 
index_labels = [
    {"label": "Sub", "index_value": 1},
    {"label": "IS K-8", "index_value": 15},
    {"label": "IS 9-12", "index_value": 2},
    {"label": "SWD K-8", "index_value": 4},
    {"label": "SWD 9-12", "index_value": 5},
    {"label": "OSS K-8", "index_value": 9},
    {"label": "OSS 9-12", "index_value": 10},
    {"label": "EX K-8", "index_value": 11},
    {"label": "EX 9-12", "index_value": 12},
    {"label": "ER", "index_value": 13},
    {"label": "MDM", "index_value": 14},
]

# Create target field list with labels and ref indexes
target_field_list = []
for i in index_labels:
    label = i["label"]
    index = i["index_value"]
    value = refined_numbers[index]

    # Create object with label and target value
    field_object = {"label": label, "value": value}
    # Add object to list 
    target_field_list.append(field_object) 

# Create school object and print data
school = School(school_name, is_both, target_field_list)
print(school.name)
print(f"ES and HS ? : {school.is_both}")
for obj in school.data_list:
     label = obj["label"]
     value = obj["value"]
     print(f"{label}: {value}")

# Return blank if value is zero
def _blank_if_zero(v):
    """Return None (blank) for 0/ '0'/ None; else return v."""
    if v is None:
        return None
    try:
        return None if int(v) == 0 else v
    except Exception:
        return v
    
# Method for retriving label values from school data list
def _get(school, label):
    for d in school.data_list:
        if d.get("label") == label:
            return d.get("value")
    return None

def build_rows_for_school(school):
    rows = []

    # pull all values
    name = school.name
    sub        = _blank_if_zero(_get(school, "Sub"))

    # K-8
    k8_students = _blank_if_zero(_get(school, "SWD K-8"))
    k8_teachers = _blank_if_zero(_get(school, "IS K-8"))
    k8_oss      = _blank_if_zero(_get(school, "OSS K-8"))
    k8_ex       = _blank_if_zero(_get(school, "EX K-8"))

    # 9-12
    hs_students = _blank_if_zero(_get(school, "SWD 9-12"))
    hs_teachers = _blank_if_zero(_get(school, "IS 9-12"))
    hs_oss      = _blank_if_zero(_get(school, "OSS 9-12"))
    hs_ex       = _blank_if_zero(_get(school, "EX 9-12"))

    er  = _blank_if_zero(_get(school, "ER"))
    mdm = _blank_if_zero(_get(school, "MDM"))

    # Verify ES or HS from values
    is_es = any(v is not None for v in [k8_students, k8_teachers, k8_oss, k8_ex])
    is_hs = any(v is not None for v in [hs_students, hs_teachers, hs_oss, hs_ex])

    # Conditional check for ES, HS, or BOTH 
    if school.is_both or (is_es and is_hs):
        rows.append({
            "School": name + " ES", 
            "Students": k8_students, "Teachers": k8_teachers, "Sub": sub,
            "OSS": k8_oss, "EX": k8_ex, "ER": er, "MDM": mdm,
        })
        rows.append({
            "School": name + " HS",
            "Students": hs_students, "Teachers": hs_teachers, "Sub": sub,
            "OSS": hs_oss, "EX": hs_ex, "ER": er, "MDM": mdm,
        })
    else:
        # Only ES data present → single ES row
        if is_es and not is_hs:
            rows.append({
                "School": name, "Level": "ES",
                "Students": k8_students, "Teachers": k8_teachers, "Sub": sub,
                "OSS": k8_oss, "EX": k8_ex, "ER": er, "MDM": mdm,
            })
        # Only HS data present → single HS row
        elif is_hs and not is_es:
            rows.append({
                "School": name, "Level": "HS",
                "Students": hs_students, "Teachers": hs_teachers, "Sub": sub,
                "OSS": hs_oss, "EX": hs_ex, "ER": er, "MDM": mdm,
            })
        else:
            # No level data — still write a single blank row for visibility
            rows.append({
                "School": name, "Level": "ES" if not school.is_both else "ES",
                "Students": None, "Teachers": None, "Sub": sub,
                "OSS": None, "EX": None, "ER": er, "MDM": mdm,
            })

    return rows

COLUMNS = ["School", "Students", "Teachers", "Sub", "OSS", "EX", "ER", "MDM"]

def append_rows_to_excel(rows, path=OUT_PATH):
    # Create file if needed then append rows
    if not path.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.append(COLUMNS)
    else:
        wb = load_workbook(path)
        ws = wb.active

    for r in rows:
        ws.append([r.get(col) for col in COLUMNS])

    wb.save(path)
    print(f"Appended {len(rows)} row(s) → {path}")

rows = build_rows_for_school(school)
append_rows_to_excel(rows)
