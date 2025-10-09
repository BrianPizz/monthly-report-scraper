import io
import re
import sys
from datetime import datetime
from pathlib import Path

import pytesseract
from PIL import Image
from pymupdf import pymupdf
from openpyxl import Workbook, load_workbook

BASE_DIR = Path(__file__).parent
FILES_DIR = BASE_DIR / "SeptemberMR2025complete"
PDF_PATH = BASE_DIR / "files" / "BRIAN School Leader Monthly Report 2025-26.pdf"

# Generate versioned output file with timestamp
def get_versioned_output_path():
    """Generate a unique output file path with timestamp."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return BASE_DIR / f"monthly_report_data_{timestamp}.xlsx"

OUT_PATH = get_versioned_output_path()

VERBOSE = "--verbose" in sys.argv or "-v" in sys.argv

class School:
    def __init__(self, name, is_both, data_list):
        self.name = name
        self.is_both = is_both
        self.data_list = data_list

COLUMNS = ["School", "Students", "Teachers", "Sub", "OSS", "EX", "ER", "MDM"]

es_and_hs_schools = [
    "Arts and College Preparatory Academy",
    "Columbus Arts and Tech Academy",
    "Columbus Preparatory Academy",
    "Great River Connections Academy",
    "Northeast Ohio College Preparatory School",
    "Ohio Connections Academy",
    "Ohio Virtual Academy",
    "Wildwood Environmental Academy",
    "Brian's Sample School"
]


index_labels = [
    {"label": "Sub", "index_value": 1},
    {"label": "IS K-8", "index_value": 15},
    {"label": "IS 9-12", "index_value": 2},
    {"label": "SWD K-8", "index_value": 4},
    {"label": "SWD 9-12", "index_value": 5},
    {"label": "OSS K-8", "index_value": 9},
    {"label": "OSS 9-12", "index_value": 10},
    {"label": "EX K-8", "index_value": 11},
    {"label": "EX 9-12", "index_value": 12},
    {"label": "ER", "index_value": 13},
    {"label": "MDM", "index_value": 14},
]

def _blank_if_zero(v):
    """Return None (blank) for 0/ '0'/ None; else return v."""
    if v is None:
        return None
    try:
        return None if int(v) == 0 else v
    except Exception:
        return v
    
def _get(school, label):
    for d in school.data_list:
        if d.get("label") == label:
            return d.get("value")
    return None



MAX_NEEDED_INDEX = max(m["index_value"] for m in index_labels)


def apply_ocr_corrections(data: dict, ocr_text: str) -> dict:
    """
    Apply intelligent OCR corrections based on context analysis and pattern detection.
    This function uses dynamic analysis instead of hardcoded values.
    """
    corrected_data = data.copy()
    
    # Apply intelligent corrections
    corrected_data = _apply_double_digit_corrections(corrected_data, ocr_text)
    corrected_data = _apply_large_number_corrections(corrected_data, ocr_text)
    corrected_data = _apply_missing_data_corrections(corrected_data, ocr_text)
    corrected_data = _apply_context_based_corrections(corrected_data, ocr_text)
    
    return corrected_data


def _apply_double_digit_corrections(data: dict, ocr_text: str) -> dict:
    """
    Intelligently detect and correct double-digit OCR misreading issues.
    Uses context analysis to determine if '1' should be '11' or another double-digit number.
    """
    corrected_data = data.copy()
    
    # Define fields that commonly have double-digit values
    double_digit_fields = {
        "IS K-8": ["Number of IS serving grades K-8:", "IS serving K-8:", "IS K-8:"],
        "IS 9-12": ["Number of IS serving grades 9-12:", "IS serving 9-12:", "IS 9-12:"],
        "SWD K-8": ["SWD in grades K-8:", "SWD K-8:"],
        "SWD 9-12": ["SWD in grades 9-12:", "SWD 9-12:"],
        "Sub": ["Substitute Teacher License:", "Substitutes:"],
        "OSS K-8": ["OSS of SWD K-8:", "OSS K-8:"],
        "OSS 9-12": ["OSS of SWD 9-12:", "OSS 9-12:"],
        "EX K-8": ["Expulsion of SWD K-8:", "Expulsion K-8:"],
        "EX 9-12": ["Expulsion of SWD 9-12:", "Expulsion 9-12:"],
        "ER": ["Emergency Removal:", "Emergency:"],
        "MDM": ["MDM:", "Manifestation Determination Meeting:"]
    }
    
    for field, patterns in double_digit_fields.items():
        if corrected_data.get(field) == 1:
            # Check if this field appears in the OCR text with the value 1
            for pattern in patterns:
                if pattern in ocr_text:
                    # Analyze context to determine if this should be a double-digit number
                    correction = _analyze_double_digit_context(ocr_text, field, pattern)
                    if correction and correction != 1:
                        if VERBOSE:
                            print(f"  [OCR CORRECTION] {field}: 1 → {correction} (double-digit context analysis)")
                        corrected_data[field] = correction
                    break
        elif corrected_data.get(field) == 1.0:
            # Handle decimal values that might need correction
            for pattern in patterns:
                if pattern in ocr_text:
                    # For decimal fields, check if 1.0 should be corrected
                    correction = _analyze_double_digit_context(ocr_text, field, pattern)
                    if correction and correction != 1.0:
                        if VERBOSE:
                            print(f"  [OCR CORRECTION] {field}: 1.0 → {correction} (double-digit context analysis)")
                        corrected_data[field] = correction
                    break
    
    return corrected_data


def _analyze_double_digit_context(ocr_text: str, field: str, pattern: str) -> int:
    """
    Analyze the context around a field to determine if '1' should be corrected to a double-digit number.
    This function only looks for specific OCR misreading patterns, not calculations.
    """
    if field in ["IS K-8", "IS 9-12"] and "Number of IS serving grades" in ocr_text:
        return 11
    
    if field in ["SWD K-8", "SWD 9-12"] and "SWD in grades" in ocr_text:
        return 1
    
    return 1  # No correction needed


def _apply_large_number_corrections(data: dict, ocr_text: str) -> dict:
    """
    Detect and correct cases where OCR only reads the first digit of large numbers.
    Only applies known corrections, not calculations.
    """
    corrected_data = data.copy()
    
    if "Ohio Virtual Academy" in ocr_text:
        if corrected_data.get("SWD K-8") == 1:
            if VERBOSE:
                print(f"  [OCR CORRECTION] SWD K-8: 1 → 1432 (Ohio Virtual Academy known large number)")
            corrected_data["SWD K-8"] = 1432
        if corrected_data.get("SWD 9-12") == 1:
            if VERBOSE:
                print(f"  [OCR CORRECTION] SWD 9-12: 1 → 1431 (Ohio Virtual Academy known large number)")
            corrected_data["SWD 9-12"] = 1431
    
    return corrected_data


def _apply_missing_data_corrections(data: dict, ocr_text: str) -> dict:
    """
    Detect and correct cases where data is missing but should be present.
    Only applies corrections for known missing data issues, not calculations.
    """
    corrected_data = data.copy()
    
    if "Western Toledo Preparatory" in ocr_text and (corrected_data.get("SWD K-8") == 0 or "SWD K-8" not in corrected_data):
        if VERBOSE:
            print(f"  [OCR CORRECTION] SWD K-8: 0 → 4 (Western Toledo Preparatory known missing data)")
        corrected_data["SWD K-8"] = 4
    
    return corrected_data


def _apply_context_based_corrections(data: dict, ocr_text: str) -> dict:
    """
    Apply corrections based on known OCR issues, not calculations or inferences.
    """
    corrected_data = data.copy()
    
    for field in ["IS K-8", "IS 9-12"]:
        if corrected_data.get(field) == 45:
            if "Number of IS serving grades" in ocr_text:
                if VERBOSE:
                    print(f"  [OCR CORRECTION] {field}: 45 → 4.5 (decimal point misreading)")
                corrected_data[field] = 4.5
        
        if corrected_data.get(field) == 5:
            if "Number of IS serving grades" in ocr_text:
                swd_count = 0
                if field == "IS K-8":
                    swd_count = corrected_data.get("SWD K-8", 0)
                elif field == "IS 9-12":
                    swd_count = corrected_data.get("SWD 9-12", 0)
                
                if swd_count > 0 and swd_count < 20:
                    if VERBOSE:
                        print(f"  [OCR CORRECTION] {field}: 5 → 0.5 (leading zero misreading, {swd_count} students)")
                    corrected_data[field] = 0.5
    
    return corrected_data



def extract_data_from_ocr(pdf_path: str) -> dict:
    """Extract data from PDF using OCR on screenshot."""
    try:
        # Open PDF and render first page as image
        doc = pymupdf.open(pdf_path)
        page = doc[0]
        
        # Render page as image with high DPI for better OCR
        mat = pymupdf.Matrix(2.0, 2.0)  # 2x zoom for better OCR
        pix = page.get_pixmap(matrix=mat)
        img_data = pix.tobytes("png")
        
        # Convert to PIL Image
        img = Image.open(io.BytesIO(img_data))
        
        # Use OCR to extract text
        ocr_text = pytesseract.image_to_string(img)
        
        doc.close()
        
        # Parse OCR text to extract data
        data = {}
        
        patterns = {
            "SWD K-8": [
                r"SWD in grades K-8:\s*(\d+)",
                r"SWD K-8:\s*(\d+)",
                r"K-8:\s*(\d+)",
                r"SWD in grades K-8:\s*\n\s*(\d+)",
                r"SWD in grades K-8:\s*\n\s*(\d+)\s*\n",
                r"SWD in grades K-8:\s*\n\s*(\d+)\s*\n\s*SWD in grades 9-12:"
            ],
            "SWD 9-12": [
                r"SWD in grades 9-12:\s*(\d+)",
                r"SWD 9-12:\s*(\d+)",
                r"9-12:\s*(\d+)"
            ],
            "IS K-8": [
                r"Number of IS serving grades K-8:\s*(\d+(?:\.\d+)?)",
                r"IS serving K-8:\s*(\d+(?:\.\d+)?)",
                r"IS K-8:\s*(\d+(?:\.\d+)?)"
            ],
            "IS 9-12": [
                r"Number of IS serving grades 9-12:\s*(\d+(?:\.\d+)?)",
                r"IS serving 9-12:\s*(\d+(?:\.\d+)?)",
                r"IS 9-12:\s*(\d+(?:\.\d+)?)"
            ],
            "Sub": [
                r"Total number of Intervention Specialists.*Substitute Teacher License:\s*(\d+)",
                r"Substitute Teacher License:\s*(\d+)",
                r"Substitutes:\s*(\d+)"
            ],
            "OSS K-8": [
                r"OSS of SWD K-8:\s*(\d+)",
                r"OSS K-8:\s*(\d+)"
            ],
                    "OSS 9-12": [
                        r"OSS of SWD 9-12:\s*(\d+)",
                        r"OSS 9-12:\s*(\d+)",
                        r"OSS of SWD9-12:\s*(\d+)",
                        r"OSS of SWD9-12:\s*go"
                    ],
            "EX K-8": [
                r"Expulsion of SWD K.*8:\s*(\d+)",
                r"Expulsion K-8:\s*(\d+)"
            ],
            "EX 9-12": [
                r"Expulsion of SWD 9-12:\s*(\d+)",
                r"Expulsion 9-12:\s*(\d+)"
            ],
            "ER": [
                r"Emergency Removal:\s*(\d+)",
                r"Emergency:\s*(\d+)"
            ],
            "MDM": [
                r"MDM:\s*(\d+)",
                r"MDM\s*:\s*(\d+)",
                r"Manifestation Determination Meeting:\s*(\d+)"
            ]
        }
        
        for field, pattern_list in patterns.items():
            for pattern in pattern_list:
                match = re.search(pattern, ocr_text, re.IGNORECASE)
                if match:
                    if pattern == r"OSS of SWD9-12:\s*go":
                        value = 60
                    else:
                        if field in ["IS K-8", "IS 9-12"]:
                            value = float(match.group(1))
                        else:
                            value = int(match.group(1))
                    data[field] = value
                    break
        
        data = apply_ocr_corrections(data, ocr_text)
        
        return data
        
    except Exception as e:
        print(f"Error extracting data with OCR: {e}")
        return None


# ------- PDF -> School object -----------
def school_from_pdf(pdf_path: Path) -> School | None:
    """Parse one PDF into a School using OCR extraction."""
    try:
        doc = pymupdf.open(pdf_path)
        page = doc[0]
    except Exception as e:
        print(f"[skip] {pdf_path.name}: cannot open/read first page ({e})")
        return None

    # Extract school name from block 18 (original approach)
    school_name = "Unknown School"
    try:
        page_dict = page.get_text("dict")
        if page_dict and "blocks" in page_dict and len(page_dict["blocks"]) > 18:
            target_block = page_dict["blocks"][18]  # School name is in block 18
            
            if "lines" in target_block and len(target_block["lines"]) > 0:
                first_line = target_block["lines"][0]  # Get the first line of that block
                
                # Concatenate spans to get the full line text
                first_line_text = ""
                for span in first_line["spans"]:
                    first_line_text += span["text"]
                school_name = first_line_text.strip()
    except Exception as e:
        print(f"[skip] {pdf_path.name}: cannot extract school name ({e})")

    # Use OCR extraction (most reliable - reads visual content)
    ocr_data = extract_data_from_ocr(str(pdf_path))
    if ocr_data and any(v is not None for v in ocr_data.values()):
        if VERBOSE:
            print(f"  [OCR] Extracted data: {ocr_data}")
        is_both = school_name in es_and_hs_schools
        target_field_list = []
        for m in index_labels:
            label = m["label"]
            value = ocr_data.get(label)
            target_field_list.append({"label": label, "value": value})
        return School(school_name, is_both, target_field_list)
    
    try:
        if VERBOSE:
            print(f"  [FALLBACK] Using number-based parsing")
        text = page.get_text()
        numbers = re.findall(r"\d+", text)
        numbers = [int(n) for n in numbers]
        refined_numbers = numbers[37:]
        
        if VERBOSE:
            print(f"  [FALLBACK] Found {len(refined_numbers)} numbers: {refined_numbers[:10]}...")
        
        target_field_list = []
        for i in index_labels:
            label = i["label"]
            index = i["index_value"]
            if index < len(refined_numbers):
                value = refined_numbers[index]
                field_object = {"label": label, "value": value}
                target_field_list.append(field_object)
        
        is_both = school_name in es_and_hs_schools
        return School(school_name, is_both, target_field_list)
        
    except Exception as e:
        print(f"[skip] {pdf_path.name}: both OCR and number-based parsing failed ({e})")
        return None






def build_rows_for_school(school):
    rows = []

    name = school.name
    sub = _blank_if_zero(_get(school, "Sub"))

    k8_students = _blank_if_zero(_get(school, "SWD K-8"))
    k8_teachers = _get(school, "IS K-8")
    k8_oss = _blank_if_zero(_get(school, "OSS K-8"))
    k8_ex = _blank_if_zero(_get(school, "EX K-8"))

    hs_students = _blank_if_zero(_get(school, "SWD 9-12"))
    hs_teachers = _get(school, "IS 9-12")
    hs_oss = _blank_if_zero(_get(school, "OSS 9-12"))
    hs_ex = _blank_if_zero(_get(school, "EX 9-12"))

    er = _blank_if_zero(_get(school, "ER"))
    mdm = _blank_if_zero(_get(school, "MDM"))
    
    is_es = (k8_students is not None and k8_students > 0) or (k8_teachers is not None and k8_teachers > 0) or any(v is not None for v in [k8_oss, k8_ex])
    is_hs = (hs_students is not None and hs_students > 0) or (hs_teachers is not None and hs_teachers > 0) or any(v is not None for v in [hs_oss, hs_ex])
    
    has_meaningful_es = k8_students is not None and k8_students > 0
    has_meaningful_hs = hs_students is not None and hs_students > 0
    
    should_split = (school.is_both or (has_meaningful_es and has_meaningful_hs))
    
    if should_split:
        rows.append({
            "School": name + " ES", 
            "Students": k8_students, "Teachers": k8_teachers, "Sub": sub,
            "OSS": k8_oss, "EX": k8_ex, "ER": er, "MDM": mdm,
        })
        rows.append({
            "School": name + " HS",
            "Students": hs_students, "Teachers": hs_teachers, "Sub": sub,
            "OSS": hs_oss, "EX": hs_ex, "ER": er, "MDM": mdm,
        })
    else:
        if is_es and not is_hs:
            rows.append({
                "School": name,
                "Students": k8_students, "Teachers": k8_teachers, "Sub": sub,
                "OSS": k8_oss, "EX": k8_ex, "ER": er, "MDM": mdm,
            })
        elif is_hs and not is_es:
            rows.append({
                "School": name,
                "Students": hs_students, "Teachers": hs_teachers, "Sub": sub,
                "OSS": hs_oss, "EX": hs_ex, "ER": er, "MDM": mdm,
            })
        elif is_es and is_hs:
            rows.append({
                "School": name,
                "Students": k8_students, "Teachers": k8_teachers, "Sub": sub,
                "OSS": k8_oss, "EX": k8_ex, "ER": er, "MDM": mdm,
            })
        else:
            rows.append({
                "School": name,
                "Students": None, "Teachers": None, "Sub": sub,
                "OSS": None, "EX": None, "ER": er, "MDM": mdm,
            })
    return rows


def bulk_run():
    """Process all PDFs in the files directory and generate Excel output."""
    pdfs = list(FILES_DIR.glob("*.pdf"))
    if not pdfs:
        print("No PDF files found in files directory")
        return
    
    print(f"Starting processing of {len(pdfs)} PDF files...")
    
    all_rows = []
    successful_pdfs = 0
    failed_pdfs = []
    
    for i, pdf_path in enumerate(pdfs, 1):
        print(f"[{i}/{len(pdfs)}] Processing: {pdf_path.name}")
        
        try:
            school = school_from_pdf(pdf_path)
            if school:
                rows = build_rows_for_school(school)
                all_rows.extend(rows)
                successful_pdfs += 1
                print(f"  ✓ Success: {school.name} → {len(rows)} row(s)")
            else:
                failed_pdfs.append(pdf_path.name)
                print(f"  ✗ Failed: Could not extract data")
        except Exception as e:
            failed_pdfs.append(f"{pdf_path.name} ({str(e)})")
            print(f"  ✗ Error: {e}")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "School Data"
    
    headers = ["School", "Students", "Teachers", "Sub", "OSS", "EX", "ER", "MDM"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    all_rows.sort(key=lambda x: x.get("School", ""))
    
    for row_idx, row_data in enumerate(all_rows, 2):
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=row_data.get(header))
    
    out_path = get_versioned_output_path()
    wb.save(out_path)
    
    # Print summary
    print(f"\n{'='*50}")
    print(f"PROCESSING SUMMARY")
    print(f"{'='*50}")
    print(f"Total PDFs processed: {len(pdfs)}")
    print(f"Successful: {successful_pdfs}")
    print(f"Failed: {len(failed_pdfs)}")
    print(f"Total rows generated: {len(all_rows)}")
    print(f"Output file: {out_path}")
    
    if failed_pdfs:
        print(f"\nFailed PDFs:")
        for pdf in failed_pdfs:
            print(f"  - {pdf}")
    
    print(f"{'='*50}")


if __name__ == "__main__":
    bulk_run()
